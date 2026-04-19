import uuid
import pyodbc
from openpyxl import load_workbook
from config import get_connection_string, get_connection_string_driver13

xlsx_path = r"e:\XTRA_WEB\مطاعم\tb_500_egyptian_restaurant_raw_materials.xlsx"

conn = None
for fn in (get_connection_string, get_connection_string_driver13):
    try:
        conn = pyodbc.connect(fn(), timeout=10)
        break
    except Exception:
        conn = None
if not conn:
    raise SystemExit("DB connection failed")

cur = conn.cursor()
cur.execute("SELECT DB_NAME()")
db_name = cur.fetchone()[0]

cur.execute("SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME='TBL007'")
tbl007_cols = {r[0] for r in cur.fetchall()}

cur.execute("SELECT CardGuide, GroupName FROM TBL006 WHERE GroupName IS NOT NULL")
group_map = {str(r[1]).strip(): str(r[0]).upper() for r in cur.fetchall() if r[0] and r[1]}

wb = load_workbook(xlsx_path, data_only=True)
ws = wb['tb']
headers = [str(ws.cell(1, c).value).strip() if ws.cell(1, c).value is not None else '' for c in range(1, ws.max_column + 1)]
idx = {h: i for i, h in enumerate(headers)}

def getv(row, name):
    i = idx.get(name)
    if i is None:
        return None
    v = row[i]
    return None if v is None else str(v).strip()

inserted = 0
updated = 0
skipped_no_group = 0
skipped_bad = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    card_code = getv(row, 'item_code')
    category = getv(row, 'category')
    ar = getv(row, 'item_name_ar')
    en = getv(row, 'item_name_en')
    is_active = (getv(row, 'is_active') or 'Yes').lower()

    if not card_code or not ar:
        skipped_bad += 1
        continue

    group_guid = group_map.get(category or '')
    if not group_guid:
        skipped_no_group += 1
        continue

    not_active = 0 if is_active in ('yes', 'y', '1', 'true', 'active') else 1

    cur.execute("SELECT TOP 1 CardGuide FROM TBL007 WHERE CardCode = ?", (card_code,))
    ex = cur.fetchone()

    if ex:
        sets, vals = [], []
        def add_set(col, val):
            if col in tbl007_cols:
                sets.append(f"[{col}] = ?")
                vals.append(val)

        add_set('ProductName', ar)
        add_set('LatinName', en)
        add_set('GroupGuid', group_guid)
        add_set('NotActive', not_active)
        add_set('StockProduct', 1)
        add_set('Security', 1)

        if sets:
            sql = f"UPDATE TBL007 SET {', '.join(sets)} WHERE CardCode = ?"
            vals.append(card_code)
            cur.execute(sql, tuple(vals))
        updated += 1
    else:
        cols, vals = [], []
        def add_col(col, val):
            if col in tbl007_cols:
                cols.append(col)
                vals.append(val)

        add_col('CardGuide', str(uuid.uuid4()).upper())
        add_col('CardCode', card_code)
        add_col('ProductName', ar)
        add_col('LatinName', en)
        add_col('GroupGuid', group_guid)
        add_col('NotActive', not_active)
        add_col('StockProduct', 1)
        add_col('Security', 1)
        add_col('ProductType', 1)
        add_col('ListAlternatives', 0)
        add_col('TaxRatio', 0)
        add_col('AgentPrice', 0)
        add_col('EndUserPrice', 0)

        if 'CardGuide' not in cols or 'ProductName' not in cols:
            skipped_bad += 1
            continue

        placeholders = ','.join(['?'] * len(cols))
        col_sql = ','.join([f'[{c}]' for c in cols])
        cur.execute(f"INSERT INTO TBL007 ({col_sql}) VALUES ({placeholders})", tuple(vals))
        inserted += 1

conn.commit()
print(f"DB={db_name}")
print(f"INSERTED={inserted}")
print(f"UPDATED={updated}")
print(f"SKIPPED_NO_GROUP={skipped_no_group}")
print(f"SKIPPED_BAD={skipped_bad}")
print(f"TOTAL_ROWS={ws.max_row-1}")
conn.close()
